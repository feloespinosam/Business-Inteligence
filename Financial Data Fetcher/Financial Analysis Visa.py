import yfinance as yf
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ─── TICKERS ──────────────────────────────────────────────────────────────────
tickers = [
    'TSLA'
]

# ─── FORMATOS EXCEL ───────────────────────────────────────────────────────────
FMT_ACCOUNTING  = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
FMT_ACCOUNTING2 = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_PERCENT     = '0.00%'
FMT_RATIO       = '0.00"x"'

# ─── RATIOS ───────────────────────────────────────────────────────────────────
financial_ratios = {
    'Current Ratio':        ('currentRatio',      FMT_RATIO),
    'Quick Ratio':          ('quickRatio',         FMT_RATIO),
    'Cash Ratio':           ('cashRatio',          FMT_RATIO),
    'Gross Margin':         ('grossMargins',       FMT_PERCENT),
    'Operating Margin':     ('operatingMargins',   FMT_PERCENT),
    'Net Profit Margin':    ('profitMargins',      FMT_PERCENT),
    'EBITDA Margin':        ('ebitdaMargins',      FMT_PERCENT),
    'ROE':                  ('returnOnEquity',     FMT_PERCENT),
    'ROA':                  ('returnOnAssets',     FMT_PERCENT),
    'ROIC':                 ('returnOnCapital',    FMT_PERCENT),
    'EPS (Trailing)':       ('trailingEps',        FMT_ACCOUNTING2),
    'EPS (Forward)':        ('forwardEps',         FMT_ACCOUNTING2),
    'Debt to Equity':       ('debtToEquity',       FMT_RATIO),
    'Total Debt to Assets': ('debtRatio',          FMT_RATIO),
    'Interest Coverage':    ('interestCoverage',   FMT_RATIO),
    'Financial Leverage':   ('financialLeverage',  FMT_RATIO),
    'Asset Turnover':       ('assetTurnover',      FMT_RATIO),
    'Inventory Turnover':   ('inventoryTurnover',  FMT_RATIO),
    'Revenue per Share':    ('revenuePerShare',     FMT_ACCOUNTING2),
    'Operating Cash Flow':  ('operatingCashflow',  FMT_ACCOUNTING),
    'Free Cash Flow':       ('freeCashflow',        FMT_ACCOUNTING),
    'Revenue Growth':       ('revenueGrowth',       FMT_PERCENT),
    'Earnings Growth':      ('earningsGrowth',      FMT_PERCENT),
}

# ─── ORDEN DE FILAS POR ESTADO FINANCIERO ─────────────────────────────────────
INCOME_ORDER = [
    # ── Ingresos ──────────────────────────────────────────────────────────────
    'Total Revenue',
    'Operating Revenue',
    'Excise Taxes',

    # ── Costos ────────────────────────────────────────────────────────────────
    'Cost Of Revenue',
    'Reconciled Cost Of Revenue',
    'Gross Profit',

    # ── Gastos Operacionales ──────────────────────────────────────────────────
    'Operating Expense',
    'Selling General And Administration',
    'General And Administrative Expense',
    'Other Gand A',
    'Selling And Marketing Expense',
    'Research And Development',
    'Depreciation Amortization Depletion Income Statement',
    'Depreciation And Amortization In Income Statement',
    'Depreciation Income Statement',
    'Amortization',
    'Amortization Of Intangibles Income Statement',
    'Reconciled Depreciation',
    'Provision For Doubtful Accounts',
    'Occupancy And Equipment',
    'Insurance And Claims',
    'Salaries And Wages',
    'Other Operating Expenses',
    'Total Expenses',

    # ── Utilidad Operacional ──────────────────────────────────────────────────
    'Operating Income',
    'Total Operating Income As Reported',
    'EBITDA',
    'Normalized EBITDA',
    'EBIT',

    # ── Ingresos / Gastos No Operacionales ───────────────────────────────────
    'Interest Income',
    'Interest Income Non Operating',
    'Interest Expense',
    'Interest Expense Non Operating',
    'Net Interest Income',
    'Net Non Operating Interest Income Expense',
    'Total Other Finance Cost',
    'Earnings From Equity Interest',
    'Earnings From Equity Interest Net Of Tax',
    'Gain On Sale Of Business',
    'Gain On Sale Of Ppe',
    'Gain On Sale Of Security',
    'Other Income Expense',
    'Other Non Operating Income Expenses',
    'Other Non Interest Expense',
    'Special Income Charges',
    'Impairment Of Capital Assets',
    'Restructuring And Mergern Acquisition',
    'Write Off',
    'Other Special Charges',
    'Other Taxes',
    'Total Unusual Items',
    'Total Unusual Items Excluding Goodwill',
    'Tax Effect Of Unusual Items',

    # ── Utilidad Antes de Impuestos ───────────────────────────────────────────
    'Pretax Income',
    'Tax Provision',
    'Tax Rate For Calcs',

    # ── Utilidad Neta ─────────────────────────────────────────────────────────
    'Net Income From Continuing Operation Net Minority Interest',
    'Net Income Continuous Operations',
    'Net Income Discontinuous Operations',
    'Net Income From Continuing And Discontinued Operation',
    'Net Income Including Noncontrolling Interests',
    'Minority Interests',
    'Net Income',
    'Preferred Stock Dividends',
    'Otherunder Preferred Stock Dividend',
    'Net Income Common Stockholders',
    'Diluted NI Availto Com Stockholders',
    'Average Dilution Earnings',
    'Normalized Income',

    # ── Por Acción ────────────────────────────────────────────────────────────
    'Basic EPS',
    'Diluted EPS',
    'Basic Average Shares',
    'Diluted Average Shares',
]

BALANCE_ORDER = [
    # ══ ACTIVOS ═══════════════════════════════════════════════════════════════

    # ── Activos Corrientes ────────────────────────────────────────────────────
    'Cash And Cash Equivalents',
    'Cash Equivalents',
    'Cash Financial',
    'Cash Cash Equivalents And Short Term Investments',
    'Cash Cash Equivalents And Federal Funds Sold',
    'Restricted Cash',
    'Other Short Term Investments',
    'Available For Sale Securities',
    'Trading Securities',
    'Held To Maturity Securities',
    'Financial Assets',
    'Financial Assets Designatedas Fair Value Through Profitor Loss Total',
    'Hedging Assets Current',
    'Receivables',
    'Accounts Receivable',
    'Gross Accounts Receivable',
    'Allowance For Doubtful Accounts Receivable',
    'Accrued Interest Receivable',
    'Notes Receivable',
    'Loans Receivable',
    'Taxes Receivable',
    'Other Receivables',
    'Duefrom Related Parties Current',
    'Receivables Adjustments Allowances',
    'Inventory',
    'Raw Materials',
    'Work In Process',
    'Finished Goods',
    'Other Inventories',
    'Inventories Adjustments Allowances',
    'Prepaid Assets',
    'Assets Held For Sale Current',
    'Current Accrued Expenses',
    'Other Current Assets',
    'Current Assets',

    # ── Activos No Corrientes ─────────────────────────────────────────────────
    'Gross PPE',
    'Land And Improvements',
    'Buildings And Improvements',
    'Machinery Furniture Equipment',
    'Construction In Progress',
    'Leases',
    'Properties',
    'Other Properties',
    'Accumulated Depreciation',
    'Net PPE',
    'Investment Properties',
    'Net Investment Properties Purchase And Sale',
    'Goodwill',
    'Other Intangible Assets',
    'Goodwill And Other Intangible Assets',
    'Investments And Advances',
    'Long Term Equity Investment',
    'Investmentin Financial Assets',
    'Investmentsin Associatesat Cost',
    'Investmentsin Joint Venturesat Cost',
    'Other Investments',
    'Non Current Accounts Receivable',
    'Non Current Note Receivables',
    'Duefrom Related Parties Non Current',
    'Non Current Deferred Assets',
    'Non Current Deferred Taxes Assets',
    'Non Current Prepaid Assets',
    'Non Current Accrued Expenses',
    'Other Non Current Assets',
    'Total Non Current Assets',

    # ── Total Activos ─────────────────────────────────────────────────────────
    'Total Assets',

    # ══ PASIVOS ═══════════════════════════════════════════════════════════════

    # ── Pasivos Corrientes ────────────────────────────────────────────────────
    'Payables',
    'Accounts Payable',
    'Payables And Accrued Expenses',
    'Current Accrued Expenses',
    'Interest Payable',
    'Dividends Payable',
    'Income Tax Payable',
    'Total Tax Payable',
    'Other Payable',
    'Dueto Related Parties Current',
    'Current Debt',
    'Current Notes Payable',
    'Line Of Credit',
    'Other Current Borrowings',
    'Current Capital Lease Obligation',
    'Current Deferred Revenue',
    'Current Deferred Liabilities',
    'Current Provisions',
    'Pensionand Other Post Retirement Benefit Plans Current',
    'Derivative Product Liabilities',
    'Other Current Liabilities',
    'Current Liabilities',

    # ── Pasivos No Corrientes ─────────────────────────────────────────────────
    'Long Term Debt',
    'Long Term Capital Lease Obligation',
    'Long Term Debt And Capital Lease Obligation',
    'Capital Lease Obligations',
    'Non Current Deferred Revenue',
    'Non Current Deferred Liabilities',
    'Non Current Deferred Taxes Liabilities',
    'Non Current Pension And Other Postretirement Benefit Plans',
    'Defined Pension Benefit',
    'Employee Benefits',
    'Long Term Provisions',
    'Tradeand Other Payables Non Current',
    'Other Non Current Liabilities',
    'Total Non Current Liabilities Net Minority Interest',

    # ── Total Pasivos ─────────────────────────────────────────────────────────
    'Total Liabilities Net Minority Interest',

    # ══ PATRIMONIO ════════════════════════════════════════════════════════════
    'Preferred Stock',
    'Common Stock',
    'Capital Stock',
    'Share Issued',
    'Additional Paid In Capital',
    'Treasury Stock',
    'Treasury Shares Number',
    'Ordinary Shares Number',
    'Retained Earnings',
    'Fixed Assets Revaluation Reserve',
    'Unrealized Gain Loss',
    'Foreign Currency Translation Adjustments',
    'Gains Losses Not Affecting Retained Earnings',
    'Minimum Pension Liabilities',
    'Other Equity Adjustments',
    'Other Equity Interest',
    'Common Stock Equity',
    'Stockholders Equity',
    'Minority Interest',
    'Total Equity Gross Minority Interest',

    # ── Métricas Derivadas ────────────────────────────────────────────────────
    'Total Capitalization',
    'Invested Capital',
    'Total Debt',
    'Net Debt',
    'Working Capital',
    'Net PPE',
    'Net Tangible Assets',
    'Tangible Book Value',
    'Total Liabilities And Stockholders Equity',
]

CASHFLOW_ORDER = [
    # ══ ACTIVIDADES DE OPERACIÓN ══════════════════════════════════════════════
    'Net Income From Continuing Operations',
    'Depreciation Amortization Depletion',
    'Depreciation And Amortization',
    'Depreciation',
    'Amortization Cash Flow',
    'Amortization Of Intangibles',
    'Amortization Of Securities',
    'Stock Based Compensation',
    'Deferred Income Tax',
    'Deferred Tax',
    'Pension And Employee Benefit Expense',
    'Earnings Losses From Equity Investments',
    'Operating Gains Losses',
    'Gain Loss On Investment Securities',
    'Gain Loss On Sale Of Business',
    'Gain Loss On Sale Of PPE',
    'Unrealized Gain Loss On Investment Securities',
    'Asset Impairment Charge',
    'Provisionand Write Offof Assets',
    'Other Non Cash Items',

    # ── Cambios en Capital de Trabajo ─────────────────────────────────────────
    'Change In Working Capital',
    'Change In Receivables',
    'Changes In Account Receivables',
    'Change In Inventory',
    'Change In Prepaid Assets',
    'Change In Other Current Assets',
    'Change In Account Payable',
    'Change In Payable',
    'Change In Payables And Accrued Expense',
    'Change In Accrued Expense',
    'Change In Income Tax Payable',
    'Change In Tax Payable',
    'Change In Other Current Liabilities',
    'Change In Other Working Capital',
    'Dividend Received Cfo',
    'Interest Received Cfo',
    'Interest Paid Cfo',
    'Taxes Refund Paid',
    'Income Tax Paid Supplemental Data',
    'Interest Paid Supplemental Data',
    'Cash From Discontinued Operating Activities',
    'Other Cash Adjustment Outside Changein Cash',
    'Cash Flow From Continuing Operating Activities',
    'Operating Cash Flow',

    # ══ ACTIVIDADES DE INVERSIÓN ══════════════════════════════════════════════
    'Purchase Of PPE',
    'Sale Of PPE',
    'Net PPE Purchase And Sale',
    'Capital Expenditure',
    'Capital Expenditure Reported',
    'Purchase Of Intangibles',
    'Sale Of Intangibles',
    'Net Intangibles Purchase And Sale',
    'Purchase Of Investment',
    'Sale Of Investment',
    'Net Investment Purchase And Sale',
    'Purchase Of Investment Properties',
    'Sale Of Investment Properties',
    'Purchase Of Business',
    'Sale Of Business',
    'Net Business Purchase And Sale',
    'Net Foreign Currency Exchange Gain Loss',
    'Dividends Received Cfi',
    'Interest Received Cfi',
    'Net Other Investing Changes',
    'Cash From Discontinued Investing Activities',
    'Cash Flow From Continuing Investing Activities',
    'Investing Cash Flow',

    # ══ ACTIVIDADES DE FINANCIAMIENTO ════════════════════════════════════════
    'Issuance Of Debt',
    'Long Term Debt Issuance',
    'Short Term Debt Issuance',
    'Net Long Term Debt Issuance',
    'Net Short Term Debt Issuance',
    'Net Issuance Payments Of Debt',
    'Repayment Of Debt',
    'Long Term Debt Payments',
    'Short Term Debt Payments',
    'Issuance Of Capital Stock',
    'Common Stock Issuance',
    'Common Stock Payments',
    'Net Common Stock Issuance',
    'Proceeds From Stock Option Exercised',
    'Repurchase Of Capital Stock',
    'Cash Dividends Paid',
    'Common Stock Dividend Paid',
    'Interest Paid Cff',
    'Net Other Financing Charges',
    'Cash Flow From Continuing Financing Activities',
    'Financing Cash Flow',

    # ══ RESUMEN ═══════════════════════════════════════════════════════════════
    'Effect Of Exchange Rate Changes',
    'Changes In Cash',
    'Other Cash Adjustment Outside Changein Cash',
    'Beginning Cash Position',
    'End Cash Position',
    'Free Cash Flow',
]

def reorder_df(df, order):
    """
    Reordena las filas del df según la lista 'order'.
    Las filas que existan en el df pero no en order
    se agregan al final para no perder datos.
    """
    existing_ordered   = [r for r in order if r in df.index]
    remaining          = [r for r in df.index if r not in order]
    return df.loc[existing_ordered + remaining]

# ─── PALETA DE COLORES ────────────────────────────────────────────────────────
COLORS = {
    'header_bg':    '1F3864',
    'header_font':  'FFFFFF',
    'section_bg':   '2E75B6',
    'section_font': 'FFFFFF',
    'positive_bg':  'E2EFDA',
    'positive_font':'375623',
    'negative_bg':  'FCE4D6',
    'negative_font':'833C00',
    'neutral_bg':   'FFF2CC',
    'neutral_font': '7F6000',
    'alt_row':      'F2F2F2',
    'border':       'BFBFBF',
}

# ─── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def make_font(hex_color='000000', bold=False, size=11, italic=False):
    return Font(color=hex_color, bold=bold, size=size, italic=italic, name='Calibri')

def make_border(style='thin'):
    s = Side(style=style, color=COLORS['border'])
    return Border(left=s, right=s, top=s, bottom=s)

def make_alignment(horizontal='left', wrap=False):
    return Alignment(horizontal=horizontal, vertical='center', wrap_text=wrap)

def apply_header_style(cell, level='main'):
    if level == 'main':
        cell.fill = make_fill(COLORS['header_bg'])
        cell.font = make_font(COLORS['header_font'], bold=True, size=13)
    elif level == 'section':
        cell.fill = make_fill(COLORS['section_bg'])
        cell.font = make_font(COLORS['section_font'], bold=True, size=11)
    elif level == 'column':
        cell.fill = make_fill(COLORS['neutral_bg'])
        cell.font = make_font(COLORS['neutral_font'], bold=True, size=10)
    cell.alignment = make_alignment('center')
    cell.border = make_border()

def apply_data_style(cell, row_idx, value=None, color_negatives=False):
    bg = COLORS['alt_row'] if row_idx % 2 == 0 else 'FFFFFF'
    cell.fill = make_fill(bg)
    cell.font = make_font()
    if color_negatives and value is not None and isinstance(value, (int, float)):
        if value > 0:
            cell.fill = make_fill(COLORS['positive_bg'])
            cell.font = make_font(COLORS['positive_font'])
        elif value < 0:
            cell.fill = make_fill(COLORS['negative_bg'])
            cell.font = make_font(COLORS['negative_font'])
    cell.border = make_border('thin')
    cell.alignment = make_alignment('right')

def add_color_scale(ws, min_row, max_row, min_col, max_col):
    from openpyxl.formatting.rule import ColorScaleRule
    rule = ColorScaleRule(
        start_type='min',      start_color='FCE4D6',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max',        end_color='E2EFDA'
    )
    rng = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    ws.conditional_formatting.add(rng, rule)

def set_column_widths(ws, num_cols, index_width=34, data_width=18):
    ws.column_dimensions['A'].width = index_width
    for i in range(2, num_cols + 2):
        ws.column_dimensions[get_column_letter(i)].width = data_width

def freeze_header(ws, row=3, col=2):
    ws.freeze_panes = f"{get_column_letter(col)}{row}"

# ─── ANÁLISIS ─────────────────────────────────────────────────────────────────
def vertical_analysis(df, base_row_keyword):
    """Cada fila como proporción de la fila base. Sin mul(100) — FMT_PERCENT lo hace."""
    base_rows = [i for i in df.index if base_row_keyword.lower() in str(i).lower()]
    if not base_rows:
        return pd.DataFrame()
    base = df.loc[base_rows[0]]
    return df.div(base).round(4)          # ← sin .mul(100)

def horizontal_absolute(df):
    if df.shape[1] < 2:
        return pd.DataFrame()
    result = pd.DataFrame(index=df.index)
    cols = df.columns.tolist()
    for i in range(len(cols) - 1):
        result[f"{cols[i]} vs {cols[i+1]}"] = df[cols[i]] - df[cols[i+1]]
    return result

def horizontal_percent(df):
    """Variación porcentual. Sin mul(100) — FMT_PERCENT lo hace."""
    if df.shape[1] < 2:
        return pd.DataFrame()
    result = pd.DataFrame(index=df.index)
    cols = df.columns.tolist()
    for i in range(len(cols) - 1):
        result[f"{cols[i]} vs {cols[i+1]}"] = (
            (df[cols[i]] - df[cols[i+1]]) / df[cols[i+1]].abs()
        ).round(4)                         # ← sin mul(100)
    return result

# ─── ESCRITURA CON FORMATO ────────────────────────────────────────────────────
def write_formatted_df(ws, df, start_row, title,
                       num_format=FMT_ACCOUNTING,
                       color_negatives=False,
                       add_color_scale_flag=False,
                       per_row_formats=None):

    title_cell = ws.cell(row=start_row, column=1, value=title)
    apply_header_style(title_cell, level='section')
    if len(df.columns) > 0:
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row,   end_column=len(df.columns) + 1
        )
    start_row += 1

    apply_header_style(ws.cell(row=start_row, column=1, value='Concepto'), level='column')
    for col_idx, col_name in enumerate(df.columns, start=2):
        apply_header_style(ws.cell(row=start_row, column=col_idx, value=col_name), level='column')
    start_row += 1

    data_start_row = start_row

    for row_idx, (idx, row) in enumerate(df.iterrows()):
        idx_cell           = ws.cell(row=start_row, column=1, value=str(idx))
        idx_cell.font      = make_font(size=10)
        idx_cell.fill      = make_fill(COLORS['alt_row'] if row_idx % 2 == 0 else 'FFFFFF')
        idx_cell.border    = make_border('thin')
        idx_cell.alignment = make_alignment('left')

        for col_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=start_row, column=col_idx, value=value)
            apply_data_style(cell, row_idx, value, color_negatives=color_negatives)
            if isinstance(value, (int, float)):
                if per_row_formats and str(idx) in per_row_formats:
                    cell.number_format = per_row_formats[str(idx)]
                else:
                    cell.number_format = num_format

        start_row += 1

    if add_color_scale_flag and start_row > data_start_row:
        add_color_scale(ws,
            min_row=data_start_row, max_row=start_row - 1,
            min_col=2,              max_col=len(df.columns) + 1
        )

    return start_row + 1

# ─── MAIN ─────────────────────────────────────────────────────────────────────
output_file = 'financial_data_tesla.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for ticker_symbol in tickers:
        try:
            ticker = yf.Ticker(ticker_symbol)

            income_statement = ticker.income_stmt
            balance_sheet    = ticker.balance_sheet
            cash_flow        = ticker.cash_flow

            try:
                info = ticker.info
            except:
                info = {}

            # Formatear fechas y reordenar filas
            if not income_statement.empty:
                income_statement.columns = [str(c)[:10] for c in income_statement.columns]
                income_statement = reorder_df(income_statement, INCOME_ORDER)

            if not balance_sheet.empty:
                balance_sheet.columns = [str(c)[:10] for c in balance_sheet.columns]
                balance_sheet = reorder_df(balance_sheet, BALANCE_ORDER)

            if not cash_flow.empty:
                cash_flow.columns = [str(c)[:10] for c in cash_flow.columns]
                cash_flow = reorder_df(cash_flow, CASHFLOW_ORDER)

            # Key Metrics
            metrics_rows    = []
            per_row_formats = {}
            for metric_name, (info_key, fmt) in financial_ratios.items():
                metrics_rows.append({'Metric': metric_name, 'Value': info.get(info_key)})
                per_row_formats[metric_name] = fmt
            metrics_df = pd.DataFrame(metrics_rows).set_index('Metric')

            # Crear hoja
            sheet_name  = ticker_symbol[:31]
            current_row = 1
            pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # Título principal
            title_cell = ws.cell(row=current_row, column=1,
                                 value=f'{ticker_symbol} — Financial Data')
            apply_header_style(title_cell, level='main')
            ws.row_dimensions[current_row].height = 28
            current_row += 2

            # Key Metrics
            current_row = write_formatted_df(
                ws, metrics_df, current_row,
                title='Key Metrics',
                per_row_formats=per_row_formats
            )

            # ── Income Statement ───────────────────────────────────────────
            if not income_statement.empty:
                current_row = write_formatted_df(
                    ws, income_statement, current_row,
                    title='Income Statement',
                    num_format=FMT_ACCOUNTING,
                    add_color_scale_flag=True
                )
                v = vertical_analysis(income_statement, 'Total Revenue')
                if not v.empty:
                    current_row = write_formatted_df(
                        ws, v, current_row,
                        title='Income Statement — Análisis Vertical (%)',
                        num_format=FMT_PERCENT
                    )
                h_abs = horizontal_absolute(income_statement)
                if not h_abs.empty:
                    current_row = write_formatted_df(
                        ws, h_abs, current_row,
                        title='Income Statement — Variación Absoluta (Δ$)',
                        num_format=FMT_ACCOUNTING,
                        color_negatives=True
                    )
                h_pct = horizontal_percent(income_statement)
                if not h_pct.empty:
                    current_row = write_formatted_df(
                        ws, h_pct, current_row,
                        title='Income Statement — Variación Porcentual (Δ%)',
                        num_format=FMT_PERCENT,
                        color_negatives=True
                    )

            # ── Balance Sheet ──────────────────────────────────────────────
            if not balance_sheet.empty:
                current_row = write_formatted_df(
                    ws, balance_sheet, current_row,
                    title='Balance Sheet',
                    num_format=FMT_ACCOUNTING
                )
                v = vertical_analysis(balance_sheet, 'Total Assets')
                if not v.empty:
                    current_row = write_formatted_df(
                        ws, v, current_row,
                        title='Balance Sheet — Análisis Vertical (%)',
                        num_format=FMT_PERCENT
                    )
                h_abs = horizontal_absolute(balance_sheet)
                if not h_abs.empty:
                    current_row = write_formatted_df(
                        ws, h_abs, current_row,
                        title='Balance Sheet — Variación Absoluta (Δ$)',
                        num_format=FMT_ACCOUNTING,
                        color_negatives=True
                    )
                h_pct = horizontal_percent(balance_sheet)
                if not h_pct.empty:
                    current_row = write_formatted_df(
                        ws, h_pct, current_row,
                        title='Balance Sheet — Variación Porcentual (Δ%)',
                        num_format=FMT_PERCENT,
                        color_negatives=True
                    )

            # ── Cash Flow ──────────────────────────────────────────────────
            if not cash_flow.empty:
                current_row = write_formatted_df(
                    ws, cash_flow, current_row,
                    title='Cash Flow Statement',
                    num_format=FMT_ACCOUNTING
                )
                v = vertical_analysis(cash_flow, 'Operating Cash Flow')
                if not v.empty:
                    current_row = write_formatted_df(
                        ws, v, current_row,
                        title='Cash Flow — Análisis Vertical (%)',
                        num_format=FMT_PERCENT
                    )
                h_abs = horizontal_absolute(cash_flow)
                if not h_abs.empty:
                    current_row = write_formatted_df(
                        ws, h_abs, current_row,
                        title='Cash Flow — Variación Absoluta (Δ$)',
                        num_format=FMT_ACCOUNTING,
                        color_negatives=True
                    )
                h_pct = horizontal_percent(cash_flow)
                if not h_pct.empty:
                    current_row = write_formatted_df(
                        ws, h_pct, current_row,
                        title='Cash Flow — Variación Porcentual (Δ%)',
                        num_format=FMT_PERCENT,
                        color_negatives=True
                    )

            # Layout final
            max_cols = max(
                len(income_statement.columns) if not income_statement.empty else 0,
                len(balance_sheet.columns)    if not balance_sheet.empty    else 0,
                len(cash_flow.columns)        if not cash_flow.empty        else 0,
                1
            )
            set_column_widths(ws, num_cols=max_cols)
            freeze_header(ws, row=3, col=2)

            print(f'{ticker_symbol} completado')

        except Exception as e:
            import traceback
            print(f'Error con {ticker_symbol}:')
            traceback.print_exc()

print(f'Archivo creado: {output_file}')